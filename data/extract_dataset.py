"""
Extraction and harmonization of BUT TC pedagogical repository, real room inventory,
and teacher workloads according to official HETD regulations (4h TP = 3h TD).
"""

import os
import sys
import json
import re
import openpyxl

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(os.path.dirname(DATA_DIR), "Export_Pilotage_Departemental_Stages_20260901_1907.xlsx")
PN_JSON_PATH = os.path.join(DATA_DIR, "referentiel_pn.json")
OUTPUT_PATH = os.path.join(DATA_DIR, "dataset_tc.json")


def calculate_hetd(cm_hours: float, td_hours: float, tp_hours: float) -> float:
    """
    Calcule le volume en Heures Équivalent TD (HETD) selon la réglementation officielle :
    - 1h CM = 1.5h TD
    - 1h TD = 1.0h TD
    - 1h TP = 0.75h TD (soit 4h TP = 3h TD)
    """
    return round((cm_hours * 1.5) + (td_hours * 1.0) + (tp_hours * 0.75), 2)


def extract_semester_from_code(code: str) -> str:
    """Extract semester (e.g. 'R1.01' -> 'S1', 'SAE 2.03' -> 'S2')."""
    match = re.search(r'[RS](\d)', code, re.IGNORECASE)
    if match:
        return f"S{match.group(1)}"
    return "S1"


def parse_hours_distribution(total_hours: int, details_text: str = "") -> dict:
    """
    Decomposes total hours into CM, TD, TP and calculates statutory HETD.
    """
    tp = 0
    td = 0
    cm = 0
    
    if details_text:
        tp_match = re.search(r'(\d+)\s*h(?:eures?)?\s*de\s*TP', details_text, re.IGNORECASE)
        if tp_match:
            tp = int(tp_match.group(1))
        
        td_match = re.search(r'(\d+)\s*h(?:eures?)?\s*de\s*TD', details_text, re.IGNORECASE)
        if td_match:
            td = int(td_match.group(1))

        cm_match = re.search(r'(\d+)\s*h(?:eures?)?\s*de\s*CM', details_text, re.IGNORECASE)
        if cm_match:
            cm = int(cm_match.group(1))
            
    remaining = max(0, total_hours - (tp + td + cm))
    if remaining > 0:
        if total_hours <= 12:
            td += remaining
        elif tp > 0 and cm == 0 and td == 0:
            td += remaining
        else:
            calc_cm = (remaining // 3)
            calc_td = remaining - calc_cm
            cm += calc_cm
            td += calc_td

    hetd = calculate_hetd(cm, td, tp)

    return {
        "CM": cm,
        "TD": td,
        "TP": tp,
        "total_heures_presentiel": total_hours,
        "total_hetd": hetd
    }


def build_dataset():
    print(f"Loading National PN Referentiel from {PN_JSON_PATH}...")
    with open(PN_JSON_PATH, "r", encoding="utf-8") as f:
        pn_data = json.load(f)

    print(f"Loading Excel file from {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # 1. Teachers extraction with statutory service quotas
    teachers = {}
    if "👥 Charge Enseignants" in wb.sheetnames:
        ws_teach = wb["👥 Charge Enseignants"]
        for row in ws_teach.iter_rows(min_row=5, values_only=True):
            if row and row[0]:
                name = str(row[0]).strip()
                # Default status: PRAG (384h) or MCF (192h)
                statut = "PRAG" if any(k in name.lower() for k in ["tabellion", "pytel", "cardinale", "millet", "jeanne"]) else "MCF"
                service_statutaire = 384 if statut == "PRAG" else 192
                
                teachers[name] = {
                    "id": f"T_{len(teachers)+1:02d}",
                    "name": name,
                    "statut": statut,
                    "service_statutaire_hetd": service_statutaire,
                    "max_hours_per_day": 6,
                    "assigned_resources": []
                }

    # 2. Resources extraction with Teacher assignments
    resources = {}
    pn_resources = {r.get("code"): r for r in pn_data.get("resources", [])}
    
    if "📚 Ressources Pedagogiques" in wb.sheetnames:
        ws_res = wb["📚 Ressources Pedagogiques"]
        for row in ws_res.iter_rows(min_row=5, values_only=True):
            if not row or not row[1]:
                continue
            code = str(row[1]).strip()
            libelle = str(row[2]).strip() if row[2] else ""
            volume = int(row[4]) if row[4] and str(row[4]).isdigit() else 20
            parcours = str(row[5]).strip() if row[5] else "Tronc Commun"
            responsable = str(row[6]).strip() if row[6] else ""
            equipe_str = str(row[7]).strip() if row[7] else ""
            
            # Parse team members
            team = []
            if equipe_str:
                for member in equipe_str.split(","):
                    clean_name = re.sub(r'\(.*?\)', '', member).strip()
                    if clean_name:
                        team.append(clean_name)
                        if clean_name not in teachers:
                            teachers[clean_name] = {
                                "id": f"T_{len(teachers)+1:02d}",
                                "name": clean_name,
                                "statut": "VACATAIRE",
                                "service_statutaire_hetd": 100,
                                "max_hours_per_day": 6,
                                "assigned_resources": []
                            }
                        if code not in teachers[clean_name]["assigned_resources"]:
                            teachers[clean_name]["assigned_resources"].append(code)

            pn_info = pn_resources.get(code, {})
            hours_details = pn_info.get("hours_details", "")
            semester = extract_semester_from_code(code)
            
            hours_split = parse_hours_distribution(volume, hours_details)

            resources[code] = {
                "code": code,
                "label": libelle or pn_info.get("label", code),
                "semester": semester,
                "parcours": parcours,
                "volume_total": volume,
                "hours_split": hours_split,
                "responsable": responsable,
                "team": team,
                "requires_computer_lab": hours_split["TP"] > 0
            }

    # 3. REAL ROOMS INVENTORY FROM IUT TC (.ICS DATA)
    rooms = [
        # Amphithéâtre
        {"id": "IUTC-amphi 3", "name": "IUTC-Amphi 3", "capacity": 150, "type": "AMPHI", "equipments": ["VIDEO", "MIC", "AMPHI"]},
        
        # Salles Informatiques (TP Info)
        {"id": "IUTC-503 i", "name": "IUTC-503 (Info)", "capacity": 20, "type": "TP_INFO", "equipments": ["COMPUTERS", "VIDEO", "INTERNET"]},
        {"id": "IUTC-506 i", "name": "IUTC-506 (Info)", "capacity": 20, "type": "TP_INFO", "equipments": ["COMPUTERS", "VIDEO", "INTERNET"]},
        {"id": "IUTC-501/502 i", "name": "IUTC-501/502 (Info)", "capacity": 24, "type": "TP_INFO", "equipments": ["COMPUTERS", "VIDEO", "INTERNET"]},
        
        # Salle Négociation & Vente
        {"id": "IUTC-524 n", "name": "IUTC-524 (Négociation & Vente)", "capacity": 20, "type": "TP_NEGO", "equipments": ["CAMERAS", "AUDIO", "VIDEO", "JEUX_DE_ROLES"]},
        
        # Salles de TD standard
        {"id": "IUTC-514", "name": "IUTC-Salle 514 (TD)", "capacity": 32, "type": "TD", "equipments": ["VIDEO", "TABLEAU"]},
        {"id": "IUTC-515", "name": "IUTC-Salle 515 (TD)", "capacity": 32, "type": "TD", "equipments": ["VIDEO", "TABLEAU"]},
        {"id": "IUTC-516", "name": "IUTC-Salle 516 (TD)", "capacity": 32, "type": "TD", "equipments": ["VIDEO", "TABLEAU"]},
        {"id": "IUTC-518", "name": "IUTC-Salle 518 (TD)", "capacity": 32, "type": "TD", "equipments": ["VIDEO", "TABLEAU"]},
        {"id": "IUTC-519", "name": "IUTC-Salle 519 (TD)", "capacity": 32, "type": "TD", "equipments": ["VIDEO", "TABLEAU"]},
        {"id": "IUTC-513", "name": "IUTC-Salle 513 (TD)", "capacity": 30, "type": "TD", "equipments": ["VIDEO", "TABLEAU"]},
        {"id": "IUTC-311", "name": "IUTC-Salle 311 (TD)", "capacity": 30, "type": "TD", "equipments": ["VIDEO", "TABLEAU"]},
        {"id": "IUTC-322", "name": "IUTC-Salle 322 (TD)", "capacity": 28, "type": "TD", "equipments": ["VIDEO", "TABLEAU"]},
        
        # Laboratoires de Langues & Pédagogie active
        {"id": "IUTC-LABO 309", "name": "IUTC-Labo Langues 309", "capacity": 20, "type": "TP_LANG", "equipments": ["HEADSETS", "AUDIO", "VIDEO"]},
        {"id": "IUTC-102 - CLAAC", "name": "IUTC-102 (CLAAC Pédagogie Active)", "capacity": 30, "type": "TD_ACTIF", "equipments": ["ECRANS_MULTIPLES", "TABLES_MODULAIRES"]}
    ]

    # 4. REAL COHORTS HIERARCHY (3 PROMOTIONS TC)
    # BUT 1: 5 TD (TD1..TD5) & 10 TP (TP1A, TP1B, TP2A, TP2B, TP3A, TP3B, TP4A, TP4B, TP5A, TP5B)
    # BUT 2: Tronc commun + 3 Parcours (BDMRC, MDEE, MMPV) en FI & FA
    # BUT 3: 3 Parcours (BDMRC, MDEE, MMPV) en FI & FA
    cohorts = [
        {
            "id": "BUT1",
            "name": "BUT 1 TC (Tronc Commun)",
            "level": "BUT1",
            "mode": "FI",
            "size": 140,
            "promo_group": "BUT1_PROMO",
            "groups_td": ["TD1", "TD2", "TD3", "TD4", "TD5"],
            "groups_tp": [
                "TP1A", "TP1B",
                "TP2A", "TP2B",
                "TP3A", "TP3B",
                "TP4A", "TP4B",
                "TP5A", "TP5B"
            ],
            "alternance_weeks": []
        },
        {
            "id": "BUT2_FI",
            "name": "BUT 2 TC (Formation Initiale)",
            "level": "BUT2",
            "mode": "FI",
            "size": 75,
            "promo_group": "TC2_FI_PROMO",
            "groups_td": ["TC2_G1_BDMRC", "TC2_G2_MDEE", "TC2_G3_MMPV"],
            "groups_tp": ["TC2_TP1A", "TC2_TP1B", "TC2_TP2A", "TC2_TP2B", "TC2_TP3A", "TC2_TP3B"],
            "alternance_weeks": []
        },
        {
            "id": "BUT2_FA",
            "name": "BUT 2 TC (Formation en Alternance)",
            "level": "BUT2",
            "mode": "FA",
            "size": 35,
            "promo_group": "TC2_FA_PROMO",
            "groups_td": ["TC2_FA_BUT2"],
            "groups_tp": ["TC2_FA_TP1", "TC2_FA_TP2"],
            "alternance_weeks": [2, 4, 6, 8, 10, 12, 14]
        },
        {
            "id": "BUT3_FI",
            "name": "BUT 3 TC (Formation Initiale)",
            "level": "BUT3",
            "mode": "FI",
            "size": 65,
            "promo_group": "TC3_FI_PROMO",
            "groups_td": ["TC3_FI_G1_BDMRC", "TC3_FI_G2_MDEE", "TC3_FI_G3_MMPV"],
            "groups_tp": ["TC3_FI_TP1A", "TC3_FI_TP1B", "TC3_FI_TP2A", "TC3_FI_TP2B", "TC3_FI_TP3A", "TC3_FI_TP3B"],
            "alternance_weeks": []
        },
        {
            "id": "BUT3_FA",
            "name": "BUT 3 TC (Formation en Alternance)",
            "level": "BUT3",
            "mode": "FA",
            "size": 35,
            "promo_group": "TC3_FA_PROMO",
            "groups_td": ["TC3_FA_G1_BDMRC", "TC3_FA_G2_MDEE", "TC3_FA_G3_MMPV"],
            "groups_tp": ["TC3_FA_TP1", "TC3_FA_TP2"],
            "alternance_weeks": [1, 3, 5, 7, 9, 11, 13]
        }
    ]

    # 5. Real Calendar Structure (Lundi au Samedi, créneaux de 1h30)
    calendar_config = {
        "weeks_per_semester": 15,
        "catchup_weeks": [8, 15],
        "days": ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"],
        "daily_slots": [
            {"id": 0, "name": "M1", "time": "08:00 - 09:30", "period": "MATIN", "duration_hours": 1.5},
            {"id": 1, "name": "M2", "time": "09:45 - 11:15", "period": "MATIN", "duration_hours": 1.5},
            {"id": 2, "name": "S1", "time": "13:30 - 15:00", "period": "APRES_MIDI", "duration_hours": 1.5},
            {"id": 3, "name": "S2", "time": "15:15 - 16:45", "period": "APRES_MIDI", "duration_hours": 1.5}
        ],
        "permanent_closures": [
            {"day": "Jeudi", "period": "APRES_MIDI", "slots": [2, 3], "reason": "Fermeture Jeudi Après-midi (Sport / Vie étudiante)"},
            {"day": "Samedi", "period": "APRES_MIDI", "slots": [2, 3], "reason": "Fermeture Samedi Après-midi"}
        ]
    }

    dataset = {
        "department": "Techniques de Commercialisation (TC)",
        "teachers": list(teachers.values()),
        "resources": list(resources.values()),
        "rooms": rooms,
        "cohorts": cohorts,
        "calendar_config": calendar_config
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(dataset, f, indent=2, ensure_ascii=False)

    print(f"✅ Successfully exported TC dataset to {OUTPUT_PATH}")
    print(f"  - Teachers extracted: {len(teachers)}")
    print(f"  - Pedagogical resources: {len(resources)}")
    print(f"  - Real Rooms defined: {len(rooms)}")
    print(f"  - Cohorts defined (3 Promos): {len(cohorts)}")


if __name__ == "__main__":
    build_dataset()
